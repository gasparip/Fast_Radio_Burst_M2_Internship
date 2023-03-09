## 3 méthodes de fonctionnement ,
## SI pas d'argument, Prend en compte toutes les observations du répertoire nenufar-pulsar
## Argument unique "FRB", Prend en compte uniquement les FRBs du répertoir nenufar-pulsar
## Arguments avec le Nom exacte de la FRB ou de l'observation ("pulsar"), Prend en compte uniquement les FRBs/Pulsars mentionnés

import numpy as np
import subprocess # bibliothèque pour executer du code bash dans du python
from openpyxl import Workbook,load_workbook # bibliothèque pour ouvrir dans un excel
import sys # biliothèque pour permettre au programme de dialoguer avec le terminal

fileNameFrbNameList="./FichierTest.txt"
codeNameFrbNameList='./BaseDonneeNomPulsarMode.sh'
excelPath = '/home/pgaspari/Documents/Test/Sample1.xlsx'

mode=""
ListArg=[]
for arg in sys.argv:
	if arg == "FRB" and len(ListArg)==1 :
		ListArg+=[arg]
		mode="FRB"	
		break    
	ListArg+=[arg]

lArg=ListArg[1:]
if len(lArg) != 0 and mode!="FRB": mode="NAME"
if len(lArg) == 0: mode = "ALL"
if mode == "" : 
	mode="ALL"
	print("Pas de bonne saisie, fonctionnement classique")

print(mode)
################### On récupère dans le fichier tests les informations fuourni par le script bash à savoir les différents nom de FRB observés ainis que le nombre d'observation
#Donneee=[[récurence,FRB],[...]...]

Check = 0
Check = subprocess.call(codeNameFrbNameList) ## Check = 1 si action du bash terminé. Pas nécessaire mais permet de prevenir en cas de problème dans le bash

if Check == 1:  ## Vérifie que le dossier est lu 
	fichier = open (fileNameFrbNameList,"r")
	lines = fichier.readlines()
	fichier.close()


else:
	print("Probleme : fichierNom non écrit")


Donnees=[]
for line in lines:
	compteur=0 

	for i in line: 
		if i == " ":
			compteur +=1
		else : 
			break
	NewLine= line[compteur:].split()
	NewLine.reverse()
	if mode == "FRB" and "FRB" == NewLine[0][0:3]:
		Donnees += [NewLine]			
	elif mode == "NAME" and NewLine[0] in lArg:
		Donnees += [NewLine]
	elif mode == "ALL":
		Donnees += [NewLine]
print(Donnees)
 
############## Recuperation des différences informations pour chaque FRB trouvé précédement à savoir 1 : date, 2 : heure, 3 temps entre échantillon, 4 nombre d'échantillon, 5 nombre de row) 

## la ListeTime regroupe toutes les données brutes, peut servir () 
ListeTime=[]    # Servira pour la mise en fichier excel

for nFRB in range(len(Donnees)):
	sommeDuration = 0 	
	#nFRB=0 #B0329+54 test
	cmd_str = "psredit /databf/nenufar-pulsar/ES05/*/*/"+str(Donnees[nFRB][0])+"*.fits | grep -E 'tsamp|nsblk|nrows|stt_date|stt_time' | colrm 1 67 > /home/pgaspari/Documents/Test/test2" + 	str(Donnees[nFRB][0]) +".txt"
	print(cmd_str)	
	subprocess.run(cmd_str, shell=True)
	fichier = open ("./Test/test2" + str(Donnees[nFRB][0]) +".txt","r")
	lines2 = fichier.readlines()
	fichier.close()
	ListeTime += [[Donnees[nFRB][0] + "\n Date "]] + [[Donnees[nFRB][0] + "\n Heure "]] +  [[Donnees[nFRB][0] + "\n TempsSample "]] + [[Donnees[nFRB][0] + "\n NombreSample "]] +  [[Donnees[nFRB][0] + "\n NombreRow "]]

	for indice_measure in range( int(Donnees[nFRB][1])):
		line_de_travail = lines2[indice_measure*5:indice_measure*5+5]

		for k in range(5): #On tourne sur les cinqs données de l'échantillon Time
		#print(lines2[p*5+k].replace("\n",""))
			ListeTime[nFRB*5+k]+=[line_de_travail[k].replace("\n","")]
			#print(ListeTime)	
		sommeDuration += float(line_de_travail[2])*float(line_de_travail[3])*float(line_de_travail[4])
		#print(sommeDuration)
	Donnees[nFRB] += [sommeDuration]
	Donnees[nFRB] += ["First Day : "+ListeTime[nFRB*5][1]+" / Last Day : "+ListeTime[nFRB*5][-1]]
	#print("Liste Time",ListeTime)

# Donnees sous la forme, ListeTime = [[NbrFRB,NameFRB,DurationFRB,FirstDay&LastDay][...]...]

print('\n',Donnees,'\n')

# Excel page creation

wb = Workbook(excelPath) ## excelPath défini au début du script
wb.save(excelPath)
wb = load_workbook(excelPath)
#sheet = wb['SheetName']
ws = wb.active
ws.title = "Main Page"
ws1 = wb.create_sheet("Row Data",1) #Création d'une seconde page avec les données brutes

 	
## Write in the Excel Page  Pour mettre les lignes de l'excel correctement , faire un CtrA sur les datas et adapter les lignes et les colonnes au texte. 

#Row Data
for i in range(1,len(ListeTime)+1):
	for j in range(1,len(ListeTime[i-1])+1):
		ws1.cell(row=j, column = i).value = ListeTime[i-1][j-1]
#Main Data
for i in range(1,len(Donnees[0])+1):
	for j in range(1,len(Donnees)+1):
		ws.cell(row=j, column = i).value = Donnees[j-1][i-1]
wb.save(excelPath) 
		


